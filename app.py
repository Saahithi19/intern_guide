from flask import Flask, render_template_string, request, redirect, url_for
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

EXCEL_FILE = 'interns.xlsx'

# Initialize Excel file if not present
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Interns"
        ws.append(["Name", "Email", "Department", "Status"])
        wb.save(EXCEL_FILE)

# Append intern data to Excel
def save_to_excel(name, email, dept, status):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Interns"]
    ws.append([name, email, dept, status])
    wb.save(EXCEL_FILE)

# In-memory display list
interns = [
    {"name": "June", "dept": "Python", "status": "Completed"},
    {"name": "Vismay", "dept": "A360", "status": "In Progress"},
    {"name": "Saahithi", "dept": "Python", "status": "Completed"},
]

# HTML Template
base_template = """
<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='UTF-8'>
    <title>Interns Guide - {{ title }}</title>
    <style>
        body {
            margin: 0; padding: 0;
            font-family: 'Segoe UI', sans-serif;
            background: linear-gradient(135deg, #e0f7fa, #fce4ec);
            color: #333;
        }
        header {
            background: #006064;
            color: white;
            padding: 20px;
            text-align: center;
            font-size: 28px;
            font-weight: bold;
        }
        nav {
            display: flex;
            background-color: #004d40;
            justify-content: center;
            flex-wrap: wrap;
        }
        nav a {
            color: white;
            padding: 14px 20px;
            text-decoration: none;
            display: block;
            transition: background 0.3s;
        }
        nav a:hover {
            background: #00796b;
        }
        section {
            padding: 30px;
            max-width: 960px;
            margin: auto;
        }
        .card {
            background: white;
            border-radius: 12px;
            padding: 25px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
            margin-top: 20px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        th, td {
            padding: 12px;
            text-align: center;
            border-bottom: 1px solid #ccc;
        }
        th {
            background-color: #b2ebf2;
        }
        input, select {
            padding: 10px;
            margin-top: 8px;
            width: 100%;
            border-radius: 6px;
            border: 1px solid #ccc;
        }
        input[type="submit"] {
            background-color: #006064;
            color: white;
            border: none;
            cursor: pointer;
            margin-top: 15px;
        }
        input[type="submit"]:hover {
            background-color: #004d40;
        }
    </style>
</head>
<body>
<header>Interns Guide</header>
<nav>
    <a href='/'>Home</a>
    <a href='/registration'>Registration</a>
    <a href='/onboarding'>Onboarding</a>
    <a href='/modules'>Modules</a>
    <a href='/schedule'>Schedule</a>
    <a href='/resources'>Resources</a>
    <a href='/assessments'>Assessments</a>
    <a href='/contact'>Contact</a>
</nav>
<section>
    <div class='card'>
        <h2>{{ title }}</h2>
        {{ content | safe }}
    </div>
</section>
</body>
</html>
"""

@app.route('/')
def home():
    total = len(interns)
    completed = sum(1 for i in interns if i['status'].lower() == "completed")
    in_progress = total - completed
    content = f"""
        <p><strong>Total Trained Interns:</strong> {total}</p>
        <p><strong>In Progress:</strong> {in_progress}</p>
        <p><strong>Completed:</strong> {completed}</p>
    """
    return render_template_string(base_template, title="Dashboard", content=content)

@app.route('/registration', methods=['GET', 'POST'])
def registration():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        dept = request.form['dept']
        interns.append({"name": name, "dept": dept, "status": "In Progress"})
        save_to_excel(name, email, dept, "In Progress")
        return redirect(url_for('home'))

    content = """
        <form method='post'>
            Name:<br><input type='text' name='name' required><br>
            Email:<br><input type='email' name='email' required><br>
            Department:<br><input type='text' name='dept' required><br>
            <input type='submit' value='Register'>
        </form>
    """
    return render_template_string(base_template, title="Register Intern", content=content)

@app.route('/onboarding')
def onboarding():
    rows = "".join(f"<tr><td>{i['name']}</td><td>{i['dept']}</td><td>{i['status']}</td></tr>" for i in interns)
    content = f"""
        <table>
            <tr><th>Name</th><th>Department</th><th>Status</th></tr>
            {rows}
        </table>
    """
    return render_template_string(base_template, title="Onboarding Status", content=content)

@app.route('/modules')
def modules():
    content = """
        <ul>
            <li><a href='https://www.automationanywhere.com/products/robotic-process-automation' target='_blank'>A360 Learning</a></li>
            <li><a href='https://www.learnpython.org/' target='_blank'>Python Learning</a></li>
        </ul>
    """
    return render_template_string(base_template, title="Learning Modules", content=content)

@app.route('/schedule')
def schedule():
    topics = [
        "Intro to A360", "A360 Interface", "Bot Creation", "Recorder & Packages",
        "Logic & Variables", "Bot Deployment", "Python Basics", "Control Flow",
        "Functions", "Python & A360"
    ]
    rows = "".join(f"<tr><td>Day {i+1}</td><td>{topic}</td></tr>" for i, topic in enumerate(topics))
    content = f"<table><tr><th>Day</th><th>Topic</th></tr>{rows}</table>"
    return render_template_string(base_template, title="10-Day Schedule", content=content)

@app.route('/resources')
def resources():
    content = "<p>Resource materials will be updated here during the training sessions.</p>"
    return render_template_string(base_template, title="Training Resources", content=content)

@app.route('/assessments')
def assessments():
    content = """
        <ol>
            <li>Shopping Cart App</li>
            <li>Customer Onboarding</li>
            <li>Salary Automation</li>
            <li>Employee Extraction</li>
            <li>Bonafide Info Extraction</li>
            <li>Multiple File Parser</li>
            <li>Lock Breaker Challenge</li>
            <li>Save the World Task</li>
            <li>Web to Excel Data Sync</li>
        </ol>
    """
    return render_template_string(base_template, title="Assessment Tasks", content=content)

@app.route('/contact')
def contact():
    content = """
        <p>Priyadharsini - 90256XXXXX</p>
        <p>Yuvaraj - 87109XXXXX</p>
        <p>Rajesh - 78965XXXXX</p>
    """
    return render_template_string(base_template, title="Contact Us", content=content)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    app.run(debug=False, host='0.0.0.0', port=port)
